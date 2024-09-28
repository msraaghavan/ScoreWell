import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import plotly.express as px

# Initialize Excel file for database
file_path = 'student_scores.xlsx'

if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Scores'
    ws.append(["Subject", "Evaluation Type", "Test/Project Name", "Weightage", "Student Marks", "Competitors Marks", "Class Average"])
    wb.save(file_path)

# Load workbook
wb = load_workbook(file_path)
ws = wb.active

# Function to load data into pandas DataFrame
def load_data():
    return pd.read_excel(file_path, sheet_name='Scores')

# Function to save a new entry into the Excel file
def save_data(subject, eval_type, eval_name, weightage, student_marks, competitor_marks, class_avg):
    ws.append([subject, eval_type, eval_name, weightage, student_marks, competitor_marks, class_avg])
    wb.save(file_path)

# Dashboard Overview
def dashboard():
    st.header("Student Performance Dashboard")
    df = load_data()
    
    if df.empty:
        st.warning("No data available yet. Add subjects and scores to see the dashboard.")
    else:
        st.write("### Summary Table")
        st.dataframe(df)

        # Bar Chart: Comparison of Student and Class Average Marks
        fig = px.bar(df, x='Test/Project Name', y=['Student Marks', 'Class Average'], color='Subject',
                     barmode='group', title="Student Marks vs Class Average")
        st.plotly_chart(fig)

        # Pie Chart: Distribution of Weightage
        weightage_fig = px.pie(df, values='Weightage', names='Test/Project Name', title="Weightage Distribution")
        st.plotly_chart(weightage_fig)

# Add Subject
def add_subject():
    subject = st.text_input("Enter Subject Name")
    if st.button("Add Subject"):
        st.session_state['subjects'].append(subject)
        st.success(f"Subject '{subject}' added successfully!")

# View All Subjects and Add Evaluations
def view_subjects():
    if not st.session_state['subjects']:
        st.warning("No subjects added yet. Please add subjects first.")
    else:
        subject = st.selectbox("Select Subject", st.session_state['subjects'])
        if subject:
            option = st.selectbox("Choose Action", ['Add Test', 'Add Project'])
            eval_name = st.text_input(f"Enter {option} Name")
            weightage = st.number_input("Enter Weightage")
            student_marks = st.number_input("Enter Your Marks")
            class_avg = st.number_input("Enter Class Average Marks")

            if st.button(f"Save {option}"):
                # Save data for student marks
                save_data(subject, option, eval_name, weightage, student_marks, '', class_avg)
                st.success(f"{option} '{eval_name}' added successfully!")
            
            # Add Competitors
            if st.button("Add Competitor Marks"):
                num_competitors = st.number_input("How many competitors?", min_value=1, step=1)
                for i in range(int(num_competitors)):
                    competitor_name = st.text_input(f"Enter Competitor {i+1} Name")
                    competitor_marks = st.number_input(f"Enter {competitor_name}'s Marks")
                    save_data(subject, option, eval_name, '', '', competitor_marks, '')
                st.success(f"Competitor marks for {eval_name} added!")

# Main Application Layout
st.title("Student Internal Score Tracker")

# Initialize session state
if 'subjects' not in st.session_state:
    st.session_state['subjects'] = []

# Navigation
menu = ["Add Subject", "View Subjects", "Dashboard"]
choice = st.sidebar.selectbox("Menu", menu)

if choice == "Add Subject":
    add_subject()
elif choice == "View Subjects":
    view_subjects()
elif choice == "Dashboard":
    dashboard()

