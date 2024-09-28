import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl import Workbook, load_workbook
import os

# Initialize Excel file for database
file_path = 'student_scores.xlsx'

if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Scores'
    ws.append(["Subject", "Evaluation Type", "Test/Project Name", "Weightage", "Student Marks", "Competitors Marks", "Class Average"])
    wb.save(file_path)

wb = load_workbook(file_path)
ws = wb.active

subjects = []

# Base class for User and Competitor
class Person:
    def __init__(self, name):
        self.name = name
        self.scores = {}

    def add_score(self, subject, test_name, score, class_avg):
        if subject not in self.scores:
            self.scores[subject] = []
        self.scores[subject].append({
            'Test/Project': test_name,
            'Score': score,
            'Class Average': class_avg
        })

class User(Person):
    pass

class Competitor(Person):
    pass

# Function to add a new subject
def add_subject():
    subject = simpledialog.askstring("Input", "Enter subject name:")
    if subject:
        subjects.append(subject)
        messagebox.showinfo("Success", f"Subject '{subject}' added successfully!")

# Function to view all subjects and choose to add test or project
def view_subjects():
    if not subjects:
        messagebox.showinfo("No Subjects", "No subjects have been added yet!")
        return

    subject_selection_window = tk.Toplevel(root)
    subject_selection_window.title("View Subjects")

    for subject in subjects:
        tk.Button(subject_selection_window, text=subject, command=lambda s=subject: subject_options(s)).pack(pady=5)

# Options for a particular subject (Add Test/Project)
def subject_options(subject):
    options_window = tk.Toplevel(root)
    options_window.title(f"{subject} Options")

    tk.Button(options_window, text="Add Test", command=lambda: add_evaluation(subject, 'Test')).pack(pady=10)
    tk.Button(options_window, text="Add Project", command=lambda: add_evaluation(subject, 'Project')).pack(pady=10)

# Function to add a test or project
def add_evaluation(subject, eval_type):
    eval_name = simpledialog.askstring("Input", f"Enter {eval_type} name (e.g., Quiz 1):")
    weightage = simpledialog.askfloat("Input", f"Enter weightage of {eval_name}:")
    student_marks = simpledialog.askfloat("Input", f"Enter your marks for {eval_name}:")
    class_avg = simpledialog.askfloat("Input", f"Enter class average marks for {eval_name}:")
    
    # Save data to Excel
    ws.append([subject, eval_type, eval_name, weightage, student_marks, '', class_avg])
    wb.save(file_path)

    # Add competitors
    add_competitors(subject, eval_name, eval_type)

    messagebox.showinfo("Success", f"{eval_type} '{eval_name}' added successfully!")

# Function to add competitors' marks for a specific test/project
def add_competitors(subject, eval_name, eval_type):
    num_competitors = simpledialog.askinteger("Input", f"How many competitors for {eval_name}?")
    
    for _ in range(num_competitors):
        competitor_name = simpledialog.askstring("Input", f"Enter competitor's name:")
        competitor_marks = simpledialog.askfloat("Input", f"Enter marks for {competitor_name}:")
        
        # Save competitor data to Excel
        ws.append([subject, eval_type, eval_name, '', '', competitor_marks, ''])
        wb.save(file_path)

# Main Tkinter window
root = tk.Tk()
root.title("Student Score Tracker")

# Buttons for main features
tk.Button(root, text="Add Subject", command=add_subject).pack(pady=10)
tk.Button(root, text="View All Subjects", command=view_subjects).pack(pady=10)

root.mainloop()
