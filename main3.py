import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
import plotly.express as px
import plotly.graph_objects as go

# Initialize Excel file for database
file_path = 'student_scores.xlsx'
competitors_path = 'competitors.xlsx'

if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Scores'
    ws.append(["Subject", "Evaluation Type", "Test/Project Name", "Weightage", "Student Marks", "Competitor", "Competitor Marks", "Class Average"])
    wb.save(file_path)

if not os.path.exists(competitors_path):
    wb_comp = Workbook()
    ws_comp = wb_comp.active
    ws_comp.title = 'Competitors'
    ws_comp.append(["Competitor Name"])
    wb_comp.save(competitors_path)

# Load workbooks
wb = load_workbook(file_path)
ws = wb.active
wb_comp = load_workbook(competitors_path)
ws_comp = wb_comp.active

# Function to load data into pandas DataFrame
def load_data():
    return pd.read_excel(file_path, sheet_name='Scores')

# Function to load competitors
def load_competitors():
    return pd.read_excel(competitors_path, sheet_name='Competitors')

# Function to save a new entry into the Excel file
def save_data(subject, eval_type, eval_name, weightage, student_marks, competitor, competitor_marks, class_avg):
    ws.append([subject, eval_type, eval_name, weightage, student_marks, competitor, competitor_marks, class_avg])
    wb.save(file_path)

# Function to save competitor profile
def save_competitor(name):
    ws_comp.append([name])
    wb_comp.save(competitors_path)

import plotly.colors as pc

def dashboard():
    st.header("Student Performance Dashboard")
    df = load_data()

    if df.empty:
        st.warning("No data available yet. Add subjects and scores to see the dashboard.")
    else:
        subject = st.selectbox("Select Subject for Analysis", df['Subject'].unique())
        if subject:
            subject_data = df[df['Subject'] == subject]

            if 'Competitor' in subject_data.columns:
                
                # User vs Competitor Analysis
                st.write(f"### User vs Competitors Analysis for {subject}")

                competitors = subject_data['Competitor'].dropna().unique()
                colors = pc.qualitative.Set3
                color_map = {competitor: color for competitor, color in zip(competitors, colors)}
                color_map['User'] = 'lightblue'

                competitor_fig = go.Figure()
                competitor_fig.add_trace(go.Bar(
                    x=subject_data['Test/Project Name'],
                    y=subject_data['Student Marks'],
                    name='User',
                    marker_color='lightblue'
                ))

                for competitor in competitors:
                    competitor_marks = subject_data[subject_data['Competitor'] == competitor]['Competitor Marks']
                    test_names = subject_data[subject_data['Competitor'] == competitor]['Test/Project Name']
                    competitor_fig.add_trace(go.Bar(
                        x=test_names,
                        y=competitor_marks,
                        name=competitor,
                        marker_color=color_map.get(competitor, 'grey')
                    ))

                competitor_fig.update_layout(
                    title=f"User vs Competitor Marks for {subject}",
                    xaxis_title="Tests/Projects",
                    yaxis_title="Marks",
                    barmode='group'
                )
                st.plotly_chart(competitor_fig)
            else:
                st.warning(f"No 'Competitor' column found for subject '{subject}'.")

            st.write(f"### Weightage Distribution for {subject}")
            weightage_fig = px.pie(subject_data, values='Weightage', names='Test/Project Name', title=f"Weightage Distribution for {subject}")
            st.plotly_chart(weightage_fig)

            st.write(f"### User vs Class Average for {subject}")
            line_fig = go.Figure()
            line_fig.add_trace(go.Scatter(
                x=subject_data['Test/Project Name'], 
                y=subject_data['Student Marks'], 
                mode='lines+markers', 
                name='User', 
                line=dict(color='lightblue')
            ))

            for competitor in competitors:
                competitor_marks = subject_data[subject_data['Competitor'] == competitor]['Competitor Marks']
                line_fig.add_trace(go.Scatter(
                    x=subject_data['Test/Project Name'],
                    y=competitor_marks,
                    mode='lines+markers',
                    name=f'{competitor}',
                    line=dict(color=color_map.get(competitor, 'grey'))
                ))

            line_fig.add_trace(go.Scatter(
                x=subject_data['Test/Project Name'], 
                y=subject_data['Class Average'], 
                mode='lines+markers', 
                name='Class Average', 
                line=dict(color='green')
            ))

            line_fig.update_layout(
                title=f"User vs Class Average for {subject}",
                xaxis_title="Tests/Projects", 
                yaxis_title="Marks"
            )
            st.plotly_chart(line_fig)

            st.write(f"### Subject-wise Performance for {subject}")
            avg_marks = subject_data['Student Marks'].mean()
            high_test = subject_data.loc[subject_data['Student Marks'].idxmax(), 'Test/Project Name']
            low_test = subject_data.loc[subject_data['Student Marks'].idxmin(), 'Test/Project Name']
            st.write(f"Average Marks in {subject}: **{avg_marks:.2f}**")
            st.write(f"Highest Scoring Test/Project: **{high_test}**")
            st.write(f"Lowest Scoring Test/Project: **{low_test}**")

def add_subject():
    subject = st.text_input("Enter Subject Name")
    if st.button("Add Subject"):
        st.session_state['subjects'].append(subject)
        st.success(f"Subject '{subject}' added successfully!")

def view_subjects():
    if not st.session_state['subjects']:
        st.warning("No subjects added yet. Please add subjects first.")
    else:
        subject = st.selectbox("Select Subject", st.session_state['subjects'])
        if subject:
            option = st.selectbox("Choose Action", ['Add Test', 'Add Project'])
            eval_name = st.text_input(f"Enter {option} Name")
            weightage = st.number_input("Enter Weightage")
            student_marks = st.number_input("Enter Your Marks")
            class_avg = st.number_input("Enter Class Average Marks")

            if st.button(f"Save {option}"):
                save_data(subject, option, eval_name, weightage, student_marks, '', '', class_avg)
                st.success(f"{option} '{eval_name}' added successfully!")

def add_competitor():
    st.header("Add Competitor")
    competitor_name = st.text_input("Enter Competitor Name")
    if st.button("Save Competitor"):
        save_competitor(competitor_name)
        st.success(f"Competitor '{competitor_name}' added successfully!")

def view_competitors():
    st.header("Add Competitor Marks")
    df = load_data()
    competitors_df = load_competitors()

    if df.empty:
        st.warning("No tests or projects added yet. Add subjects and evaluations first.")
    elif competitors_df.empty:
        st.warning("No competitors added yet. Add competitors first.")
    else:
        eval_name = st.selectbox("Select Test/Project", df['Test/Project Name'].unique())
        competitor = st.selectbox("Select Competitor", competitors_df['Competitor Name'])
        competitor_marks = st.number_input("Enter Competitor Marks")

        if st.button("Save Competitor Marks"):
            idx = df[df['Test/Project Name'] == eval_name].index[0]
            ws.cell(row=idx + 2, column=6, value=competitor)
            ws.cell(row=idx + 2, column=7, value=competitor_marks)  # Update competitor marks in Excel
            wb.save(file_path)
            st.success(f"Competitor '{competitor}' marks for '{eval_name}' updated successfully!")

# Main Application Layout
st.title("Student Internal Score Tracker")

# Initialize session state
if 'subjects' not in st.session_state:
    st.session_state['subjects'] = []

# Navigation
menu = ["Add Subject", "View Subjects", "Add Competitor", "View Competitors", "Dashboard"]
choice = st.sidebar.selectbox("Menu", menu)

if choice == "Add Subject":
    add_subject()
elif choice == "View Subjects":
    view_subjects()
elif choice == "Add Competitor":
    add_competitor()
elif choice == "View Competitors":
    view_competitors()
elif choice == "Dashboard":
    dashboard()
